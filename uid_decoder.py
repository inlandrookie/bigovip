from openpyxl import load_workbook
wb = load_workbook(filename = 'UID.xlsx')
sheet_ranges = wb['UID']

uid_count = 1094

rev = 0

code_book  = ['B','I','G','O',
              'U','S','A','L',
              'T','N']


def reverse_num(num):
    rev = 0
    while num > 0:
        rev = (10 * rev) + num % 10
        num //= 10
    return rev

def decode(num):

    new_str = ""

    num_list = list(str(num))

    for c in map(int, num_list):

        new_str = new_str + code_book[c]

    return new_str




for row in range(2, uid_count):

    uid = sheet_ranges.cell(row=row, column=1).value

    code = reverse_num(uid)

    token = decode(code)

    sheet_ranges.cell(row=row, column=2).value = token


    #sheet_ranges.cell(row=row, column=14).value = host_share
    #sheet_ranges.cell(row=row, column=15).value = agent_share

wb.save('UID.xlsx')